#coding=utf-8
import  xlrd
from openpyxl import  load_workbook


class Excel_file(object):

    def open_Excel(self,OpFile):
        try:
            data = xlrd.open_workbook(OpFile)
            return data
        except Exception, e:
            print str(e)

    #获取excel指定sheet的总列数、行数
    def Get_FileInf(self,FileName,sheetname):
        ExcelData=self.open_Excel(FileName)
        table=ExcelData.sheet_by_name(sheetname)
        rows=table.nrows
        cols=table.ncols
        return rows,cols

    #把注册信息写入指定excel文档
    def Output_NewAccount(self,CusType,PhoneNum,Password,Name,Email,BankId,BankNo,OutPut_FileName):
        wb_account = load_workbook(OutPut_FileName)
        ws = wb_account.worksheets[0]
        OutPut_rows = ws.max_row
        OutPut_columns = ws.max_column
        print ws.title
        print   "Work Sheet Rows:", OutPut_rows
        print   "Work Sheet Cols:", OutPut_columns
        ws.cell(row=(OutPut_rows + 1), column=1).value = CusType
        ws.cell(row=(OutPut_rows + 1), column=2).value = PhoneNum
        ws.cell(row=(OutPut_rows + 1), column=3).value = Password
        ws.cell(row=(OutPut_rows + 1), column=4).value = Name
        ws.cell(row=(OutPut_rows + 1), column=5).value = Email
        ws.cell(row=(OutPut_rows + 1), column=6).value = BankNo
        ws.cell(row=(OutPut_rows + 1), column=7).value = BankId
        wb_account.save(OutPut_FileName)







