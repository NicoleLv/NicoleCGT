#coding=utf-8
from appium import webdriver
# from appium.webdriver.common.touch_action import TouchAction
# from CustomerReg.PO.base_page import BasePage
import  time
import ConfigParser
import os
from Basic_oper.Excel_oper import Excel_file
from CustomerReg.UFO_RegFunction.NewReg import NewReg
from UFO_RegFunction.HFKH import Cus_HFKH
from CustomerReg.Open_App import Install_App
from CustomerReg.Get_file import Get_File
import openpyxl

# # 通过脚本当前路径拼接其他文件的路径函数
# PATH = lambda p: os.path.abspath(os.path.join(os.path.dirname(__file__), p))
#
# cf=ConfigParser.ConfigParser()
# cf.read('Reg_Config')
# #从外部excel文档读取待注册数据的条数
# RegFile=PATH(cf.get('TestData','DataFile_Path'))
# RegSheetName=cf.get('TestData','SheetName')
# AccountToReg=Excel_file()
# rows=(AccountToReg.Get_FileInf(RegFile,RegSheetName))[0]
# cols=(AccountToReg.Get_FileInf(RegFile,RegSheetName))[1]
#
#
# #获取bankcode,districtcode文件的路径，生成随机银行卡和身份证号
# #Output_path存放注册完成的账号信息以及注册失败的截图存放路径
# BankCode_filePath=PATH(cf.get('TestFile','BankFile'))
# DistrictcodeFile_Path=PATH(cf.get('TestFile','DistributeFile'))
# OutPut_path=PATH(cf.get('TestFile','OutPutFile'))
# Screenshot_path=PATH(cf.get('TestFile','OutPutFile'))

#获取测试相关文件和数据
Get_File()
AccountToReg=Excel_file()
rows=(AccountToReg.Get_FileInf(RegFile,RegSheetName))[0]
cols=(AccountToReg.Get_FileInf(RegFile,RegSheetName))[1]

#安装待测APP并打开
driver=Install_App()



# 获取客户类型,调用不同的方法注册
for i in range(rows-1):
    TestData_Table=AccountToReg.open_Excel(RegFile).sheet_by_name(RegSheetName)
    CusType=TestData_Table.cell(i+1,0).value
    PhoneNum=TestData_Table.cell(i+1,1).value
    PassWord=TestData_Table.cell(i+1,2).value
    print 'lll'

    #根据注册账号给出的客户类型，调用不同的注册方法并开户
    if CusType==u'新用户':
        NewReg(driver,PhoneNum, PassWord, DistrictcodeFile_Path, Screenshot_path)
        #个人身份证新用户注册完开户设置交易密码
        KH_NewUser=Cus_HFKH
        KH_NewUser.Cus_HFKH(PhoneNum)
        KH_NewUser.Set_TransPassword(PassWord)
    #注册完成




